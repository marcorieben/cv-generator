"""
Module description

Purpose: analyzed as source_code
Expected Lifetime: permanent
Category: SOURCE_CODE
Created: 2026-01-23
Last Updated: 2026-01-24
"""
from datetime import datetime
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scripts.cleanup.models import FileAnalysis, CleanupReport, DecisionType




def generate_xlsx_report(report: CleanupReport) -> str:
    """
    Generate Excel report with multiple sheets for cleanup analysis.
    
    Args:
        report: CleanupReport object
        
    Returns:
        Path to generated XLSX file (temporary path)
    """
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    delete_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Orange
    keep_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # Green
    review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ===== SUMMARY SHEET =====
    summary_sheet = wb.create_sheet("Summary", 0)
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15
    
    row = 1
    summary_sheet[f'A{row}'] = "Cleanup Analysis Report"
    summary_sheet[f'A{row}'].font = Font(bold=True, size=14)
    row += 2
    
    summary_sheet[f'A{row}'] = "Run ID:"
    summary_sheet[f'B{row}'] = report.run_id
    row += 1
    
    summary_sheet[f'A{row}'] = "Mode:"
    summary_sheet[f'B{row}'] = report.mode.upper()
    row += 1
    
    summary_sheet[f'A{row}'] = "Timestamp:"
    summary_sheet[f'B{row}'] = report.timestamp
    row += 1
    
    summary_sheet[f'A{row}'] = "Total Files Analyzed:"
    summary_sheet[f'B{row}'] = report.total_files
    row += 2
    
    # Summary table
    summary_sheet[f'A{row}'] = "Decision"
    summary_sheet[f'B{row}'] = "Count"
    for col in ['A', 'B']:
        summary_sheet[f'{col}{row}'].fill = header_fill
        summary_sheet[f'{col}{row}'].font = header_font
        summary_sheet[f'{col}{row}'].border = border
    row += 1
    
    summary = report.summary
    summary_sheet[f'A{row}'] = "DELETE_SAFE"
    summary_sheet[f'B{row}'] = summary['delete_safe']
    summary_sheet[f'A{row}'].fill = delete_fill
    summary_sheet[f'B{row}'].fill = delete_fill
    row += 1
    
    summary_sheet[f'A{row}'] = "KEEP_REQUIRED"
    summary_sheet[f'B{row}'] = summary['keep_required']
    summary_sheet[f'A{row}'].fill = keep_fill
    summary_sheet[f'B{row}'].fill = keep_fill
    row += 1
    
    summary_sheet[f'A{row}'] = "REVIEW_REQUIRED"
    summary_sheet[f'B{row}'] = summary['review_required']
    summary_sheet[f'A{row}'].fill = review_fill
    summary_sheet[f'B{row}'].fill = review_fill
    
    # ===== ALL FILES SHEET =====
    files_sheet = wb.create_sheet("All Files", 1)
    
    headers = [
        "File Path",
        "Category",
        "Decision",
        "Confidence %",
        "Size (KB)",
        "Last Modified",
        "Purpose",
        "Lifetime",
        "Created",
        "Last Updated",
        "Direct Refs",
        "Indirect Refs",
        "Reasoning",
        "Risk Assessment",
        "Recommended Action"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = files_sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Set column widths
    files_sheet.column_dimensions['A'].width = 35
    files_sheet.column_dimensions['B'].width = 18
    files_sheet.column_dimensions['C'].width = 15
    files_sheet.column_dimensions['D'].width = 13
    files_sheet.column_dimensions['E'].width = 12
    files_sheet.column_dimensions['F'].width = 18
    files_sheet.column_dimensions['G'].width = 25
    files_sheet.column_dimensions['H'].width = 12
    files_sheet.column_dimensions['I'].width = 12
    files_sheet.column_dimensions['J'].width = 12
    files_sheet.column_dimensions['K'].width = 40
    files_sheet.column_dimensions['L'].width = 40
    files_sheet.column_dimensions['M'].width = 40
    
    row = 2
    for file_analysis in report.files:
        confidence_pct = int(file_analysis.confidence * 100)
        reasoning = " | ".join(file_analysis.reasoning) if file_analysis.reasoning else ""
        direct_refs = len([r for r in file_analysis.references if not r.is_indirect])
        indirect_refs = len([r for r in file_analysis.references if r.is_indirect])
        
        # Determine row color based on decision
        if file_analysis.decision == DecisionType.DELETE_SAFE:
            row_fill = delete_fill
        elif file_analysis.decision == DecisionType.KEEP_REQUIRED:
            row_fill = keep_fill
        else:
            row_fill = review_fill
        
        values = [
            file_analysis.file_path,
            file_analysis.category.value,
            file_analysis.decision.value,
            confidence_pct,
            round(file_analysis.size_kb, 1),
            file_analysis.last_modified,
            file_analysis.file_purpose,
            file_analysis.expected_lifetime,
            file_analysis.created_date,
            file_analysis.last_updated_date,
            direct_refs,
            indirect_refs,
            reasoning,
            file_analysis.risk_assessment,
            file_analysis.recommended_action
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = files_sheet.cell(row=row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        row += 1
    
    # Freeze header row
    files_sheet.freeze_panes = "A2"
    
    # ===== DELETE_SAFE SHEET =====
    delete_files = [f for f in report.files if f.decision == DecisionType.DELETE_SAFE]
    if delete_files:
        delete_sheet = wb.create_sheet("DELETE_SAFE", 2)
        delete_sheet.column_dimensions['A'].width = 35
        delete_sheet.column_dimensions['B'].width = 18
        delete_sheet.column_dimensions['C'].width = 13
        delete_sheet.column_dimensions['D'].width = 12
        
        headers = ["File Path", "Category", "Confidence %", "Size (KB)"]
        for col_idx, header in enumerate(headers, 1):
            cell = delete_sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = delete_fill
            cell.font = header_font
            cell.border = border
        
        for idx, file_analysis in enumerate(delete_files, 2):
            delete_sheet[f'A{idx}'] = file_analysis.file_path
            delete_sheet[f'B{idx}'] = file_analysis.category.value
            delete_sheet[f'C{idx}'] = int(file_analysis.confidence * 100)
            delete_sheet[f'D{idx}'] = round(file_analysis.size_kb, 1)
            
            for col in ['A', 'B', 'C', 'D']:
                delete_sheet[f'{col}{idx}'].border = border
    
    # ===== KEEP_REQUIRED SHEET =====
    keep_files = [f for f in report.files if f.decision == DecisionType.KEEP_REQUIRED]
    if keep_files:
        keep_sheet = wb.create_sheet("KEEP_REQUIRED", 3)
        keep_sheet.column_dimensions['A'].width = 35
        keep_sheet.column_dimensions['B'].width = 18
        keep_sheet.column_dimensions['C'].width = 13
        keep_sheet.column_dimensions['D'].width = 12
        
        headers = ["File Path", "Category", "Confidence %", "Size (KB)"]
        for col_idx, header in enumerate(headers, 1):
            cell = keep_sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = keep_fill
            cell.font = header_font
            cell.border = border
        
        for idx, file_analysis in enumerate(keep_files, 2):
            keep_sheet[f'A{idx}'] = file_analysis.file_path
            keep_sheet[f'B{idx}'] = file_analysis.category.value
            keep_sheet[f'C{idx}'] = int(file_analysis.confidence * 100)
            keep_sheet[f'D{idx}'] = round(file_analysis.size_kb, 1)
            
            for col in ['A', 'B', 'C', 'D']:
                keep_sheet[f'{col}{idx}'].border = border
    
    # ===== REVIEW_REQUIRED SHEET =====
    review_files = [f for f in report.files if f.decision == DecisionType.REVIEW_REQUIRED]
    if review_files:
        review_sheet = wb.create_sheet("REVIEW_REQUIRED", 4)
        review_sheet.column_dimensions['A'].width = 35
        review_sheet.column_dimensions['B'].width = 18
        review_sheet.column_dimensions['C'].width = 13
        review_sheet.column_dimensions['D'].width = 12
        review_sheet.column_dimensions['E'].width = 40
        
        headers = ["File Path", "Category", "Confidence %", "Size (KB)", "Risk Assessment"]
        for col_idx, header in enumerate(headers, 1):
            cell = review_sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = review_fill
            cell.font = header_font
            cell.border = border
        
        for idx, file_analysis in enumerate(review_files, 2):
            review_sheet[f'A{idx}'] = file_analysis.file_path
            review_sheet[f'B{idx}'] = file_analysis.category.value
            review_sheet[f'C{idx}'] = int(file_analysis.confidence * 100)
            review_sheet[f'D{idx}'] = round(file_analysis.size_kb, 1)
            review_sheet[f'E{idx}'] = file_analysis.risk_assessment
            
            review_sheet[f'E{idx}'].alignment = Alignment(wrap_text=True, vertical='top')
            for col in ['A', 'B', 'C', 'D', 'E']:
                review_sheet[f'{col}{idx}'].border = border
        
        review_sheet.freeze_panes = "A2"
    
    return wb


def save_reports(
    report: CleanupReport,
    run_folder: str,
) -> None:
    """
    Save XLSX report to disk.
    
    Args:
        report: CleanupReport object
        run_folder: Folder path to save reports to
    """
    import os
    
    # Ensure folder exists
    os.makedirs(run_folder, exist_ok=True)
    
    # Generate and save XLSX report
    xlsx_path = os.path.join(run_folder, "cleanup_report.xlsx")
    wb = generate_xlsx_report(report)
    wb.save(xlsx_path)
    
    print(f"✅ Report saved to: {xlsx_path}")
    print(f"   - Summary sheet with statistics")
    print(f"   - All Files sheet with complete details")
    print(f"   - DELETE_SAFE sheet (if any)")
    print(f"   - KEEP_REQUIRED sheet")
    print(f"   - REVIEW_REQUIRED sheet")
